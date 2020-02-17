# pip3 install pandas 
# pip3 install openpyxl
# pip3 install lxml
# pip3 install yfinance --upgrade --no-cache-dir
# pip3 install Jinja2
# pip3 install newsapi-python
# pip3 install styleframe

import pandas
from StyleFrame import StyleFrame
from pandas import DataFrame
import yfinance as yf #https://github.com/ranaroussi/yfinance
import os
from newsapi import NewsApiClient

#Please add your own key by Signing up to: https://newsapi.org/
newsapi = NewsApiClient(api_key='')


# --------------------------------------------------------------------------------- #
#This function allows you to append data
# --------------------------------------------------------------------------------- #

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

# --------------------------------------------------------------------------------- #



# --------------------------------------------------------------------------------- #
#This function colour negative percentage as red
# --------------------------------------------------------------------------------- #
def color_negative_red(val):
    color = 'red' if val < 0 else 'black'
    return 'color: %s' % color
# --------------------------------------------------------------------------------- #



#Title of the Excel file
Title_of_Excel_file = 'COMM_Lockheed_Martin'

#Range of dates you want info for about the stock
Start_date = '2020-01-25'
End_date = '2020-03-15'

#Desired Stock. NOTE:Change the Stock Symbol and name to get the info for your desired stock eg. AAPL for stock_symbol for Apple for stock_name
stock_symbol = 'LMT'
stock_name = 'Lockheed Martin'

#Delete any previous files before running the code
try:
    os.remove('./Desktop/' + Title_of_Excel_file  + '.xlsx')
except OSError:
    pass


#Title in Excel
title = []

title_excel = DataFrame({stock_name : title})

LMT = yf.Ticker(stock_symbol)

#Extracts the divident info based on range of data using Yahoo Finance
dividends = []
dividends_temp = str(LMT.history(start=Start_date, end=End_date)['Dividends']).replace("    ", ",").replace("\n", ",").replace("Date", "").replace("Name: Dividends, dtype: int64", "")[1 :-1: ].split(",")
for i in range(len(dividends_temp)):
    if (i%2 != 0):
        dividends.append(dividends_temp[i])

#Extracts the closing price of the stock
date = []
closing_price = []
data = yf.download(stock_symbol, start=Start_date, end=End_date) 
temp = str(data['Close']).replace("    ", ",").replace("\n", ",").replace("Date", "").replace("Name: Close, dtype: float64", "")[1 :-1: ].split(",")

for i in range(len(temp)):
    if (i%2 == 0):
        date.append(temp[i])
    else:
        closing_price.append(float(temp[i]))

#Calculates the percentage change in price for each day
percentage_change = pandas.Series(closing_price).pct_change()*100

News_one = []
News_two = []
News_three = []
News_four = []

#Extracts 4 news articles for a given date (all in range) for your desired stock
for i in range(len(date)):
    all_articles = newsapi.get_everything(q=stock_name,
                                      from_param=date[i],
                                      to=date[i],
                                      language='en',
                                      sort_by='popularity')

    News_one.append(all_articles['articles'][0]['source']['name'] + ": " + all_articles['articles'][0]['title'] + "    URL: " + all_articles['articles'][0]['url'])
    News_two.append(all_articles['articles'][1]['source']['name'] + ": " + all_articles['articles'][1]['title'] + "    URL: " + all_articles['articles'][1]['url'])
    News_three.append(all_articles['articles'][2]['source']['name'] + ": " + all_articles['articles'][2]['title'] + "    URL: " + all_articles['articles'][2]['url'])
    News_four.append(all_articles['articles'][3]['source']['name'] + ": " + all_articles['articles'][3]['title'] + "    URL: " + all_articles['articles'][3]['url'])


df = DataFrame({'Date': date,'Closing Price': closing_price, '% Change in Closing Price': percentage_change, 'Dividends ': dividends, 'News #1': News_one, 'News #2': News_two, 'News #3': News_three,  'News #4': News_four})
df.style.applymap(color_negative_red)

# Excel Sheet is saved in your desktop
append_df_to_excel('./Desktop/' + Title_of_Excel_file  + '.xlsx', title_excel , sheet_name='sheet1', index=False)
append_df_to_excel('./Desktop/' + Title_of_Excel_file  + '.xlsx', df, sheet_name='sheet1', index=True, startrow=1)





